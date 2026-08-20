from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import date, datetime
from io import StringIO
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.db.session import TenantSession
from app.models.corrective_action import CorrectiveAction
from app.models.data_import import DataImportJob, DataImportRow
from app.models.department import Department
from app.models.hazard import Hazard
from app.models.incident import Incident
from app.models.notification import Notification
from app.models.organisation import Organisation, OrganisationSettings
from app.models.site import Site
from app.models.sio import SIOActivity, SafetyImprovementObservation
from app.models.user import User
from app.schemas.data_import import ImportConfirmRequest
from app.schemas.sio import SIOCreate, SIORead
from app.services.data_import_service import (
    YALELO_SIO_COLUMNS,
    YALELO_SOURCE_SYSTEM,
    confirm_yalelo_sio_import,
    preview_yalelo_sio_import,
)
from app.services.export_service import export_sios_csv
from app.services.sio_service import create_sio, get_sio_analytics, list_sios
from app.services.tenancy import set_tenant_context


MAPPED_DEPARTMENTS = (
    "Administration",
    "Commercial Operations",
    "Commercial Services",
    "Engineering",
    "Environmental & Social",
    "Executive",
    "Farm Services",
    "Finance",
    "Hatchery",
    "Health & Safety",
    "Human Resource",
    "IT",
    "Lake OPS",
    "Logistic & Distribution",
    "Management",
    "Marketing",
    "Processing Plant",
    "Procurement",
    "Quality",
    "Quality & Assurance",
    "Sales",
    "Security",
    "Warehouse",
)

EXACT_USER_NAMES = (
    "Weston Musonda",
    "Esther Phiri",
    "Daniel Chisanga",
    "Lombe Mulilo",
)

FIELD_DESTINATIONS = {
    "ID": "external_reference_id",
    "Date": "observation_date",
    "Department": "department",
    "Source of Observation": "source_type",
    "Description of SIO": "description",
    "Incident Classification": "incident_classification",
    "Status": "legacy_metadata.source_status",
    "Nature of Observation": "legacy_metadata.source_observation_nature",
    "Department Responsible for Corrective Action": "responsible_department",
    "Site": "site.name",
    "Responsible H&S Officer": "responsible_hs_officer_name",
    "Urgency": "legacy_metadata.source_urgency",
    "SIO Category": "category",
    "Person Responsible for Corrective Action": "responsible_person_name",
    "Created": "source_created_at",
    "Property Damage": "property_damage",
    "Created By": "source_created_by",
    "Modified By": "source_modified_by",
    "MonthY": "legacy_metadata.MonthY",
    "Item Type": "legacy_metadata.Item Type",
    "Path": "source_path",
}


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def source_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def source_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [source_text(value) for value in next(rows)]
        return [
            {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header is not None
            }
            for values in rows
            if any(not is_blank(value) for value in values)
        ]
    finally:
        workbook.close()


def code_for(value: str) -> str:
    code = re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")
    return code[:45] or "VALUE"


def seed_tenant(
    db: Session,
    *,
    organisation_id: int,
    organisation_name: str,
    organisation_code: str,
    sites: tuple[str, ...],
    departments: tuple[str, ...] = (),
    users: tuple[str, ...] = (),
) -> dict[str, Any]:
    organisation = Organisation(
        id=organisation_id,
        name=organisation_name,
        code=organisation_code,
        slug=organisation_code.lower(),
        timezone="Africa/Lusaka",
        is_active=True,
    )
    db.add(organisation)
    db.flush()
    set_tenant_context(db, organisation_id, platform_admin=True)
    db.add(OrganisationSettings())
    site_records = [Site(name=name, code=code_for(name)) for name in sites]
    db.add_all(site_records)
    db.flush()
    department_records = [
        Department(name=name, code=code_for(name)) for name in departments
    ]
    db.add_all(department_records)
    db.flush()
    user_records = [
        User(
            email=f"audit-{organisation_code.lower()}-{index}@example.test",
            full_name=name,
            hashed_password="not-used",
            is_active=True,
            assigned_site_id=site_records[0].id,
        )
        for index, name in enumerate(users, start=1)
    ]
    actor = User(
        email=f"audit-{organisation_code.lower()}-administrator@example.test",
        full_name=f"{organisation_name} Audit Administrator",
        hashed_password="not-used",
        is_active=True,
        is_platform_admin=True,
        assigned_site_id=site_records[0].id,
    )
    db.add_all([*user_records, actor])
    db.commit()
    return {
        "organisation": organisation,
        "actor": actor,
        "sites": site_records,
        "departments": department_records,
        "users": user_records,
    }


COUNT_MODELS = {
    "organisations": Organisation,
    "sites": Site,
    "departments": Department,
    "users": User,
    "sios": SafetyImprovementObservation,
    "actions": CorrectiveAction,
    "incidents": Incident,
    "hazards": Hazard,
    "notifications": Notification,
    "import_jobs": DataImportJob,
    "import_rows": DataImportRow,
    "sio_activities": SIOActivity,
}


def counts(db: Session) -> dict[str, int]:
    return {
        name: int(db.scalar(select(func.count()).select_from(model)) or 0)
        for name, model in COUNT_MODELS.items()
    }


def normalized_compare(source: Any, destination: Any, *, field: str) -> bool:
    if is_blank(source):
        return destination is None or destination == ""
    if field == "Date":
        source_date = source.date() if isinstance(source, datetime) else source
        return destination == source_date
    if field == "Created":
        if not isinstance(source, (date, datetime)) or destination is None:
            return False
        source_datetime = (
            source
            if isinstance(source, datetime)
            else datetime.combine(source, datetime.min.time())
        )
        return destination.replace(tzinfo=None) == source_datetime.replace(tzinfo=None)
    if field == "Description of SIO":
        return destination == source
    return source_text(destination) == source_text(source)


def destination_value(record: SafetyImprovementObservation, field: str) -> Any:
    if field == "Status":
        return (record.legacy_metadata or {}).get("source_status")
    if field == "Nature of Observation":
        return (record.legacy_metadata or {}).get("source_observation_nature")
    if field == "Urgency":
        return (record.legacy_metadata or {}).get("source_urgency")
    if field == "MonthY":
        return (record.legacy_metadata or {}).get("MonthY")
    if field == "Item Type":
        return (record.legacy_metadata or {}).get("Item Type")
    if field == "Site":
        return record.site.name if record.site else None
    return getattr(record, FIELD_DESTINATIONS[field])


def data_loss_profile(
    rows: list[dict[str, Any]],
    record_lookup: dict[str, SafetyImprovementObservation],
) -> list[dict[str, Any]]:
    profile: list[dict[str, Any]] = []
    for field in YALELO_SIO_COLUMNS:
        populated = preserved = 0
        for source_row in rows:
            source = source_row.get(field)
            if is_blank(source):
                continue
            populated += 1
            external_id = source_text(source_row.get("ID"))
            record = record_lookup.get(external_id or "")
            if record is not None and normalized_compare(
                source,
                destination_value(record, field),
                field=field,
            ):
                preserved += 1
        profile.append(
            {
                "source_column": field,
                "source_populated_count": populated,
                "successfully_preserved_count": preserved,
                "intentionally_ignored_count": 0,
                "failed_or_lost_count": populated - preserved,
                "database_destination": FIELD_DESTINATIONS[field],
            }
        )
    return profile


def representative_ids(rows: list[dict[str, Any]]) -> list[str]:
    selectors: tuple[Callable[[dict[str, Any]], bool], ...] = (
        lambda row: str(row.get("Nature of Observation", "")).startswith("Positive"),
        lambda row: str(row.get("Nature of Observation", "")).startswith("Negative"),
        lambda row: is_blank(row.get("Date")),
        lambda row: not is_blank(row.get("Date")),
        lambda row: source_text(row.get("Urgency")) in {"High", "Urgent"},
        lambda row: source_text(row.get("Status")) == "Complete",
        lambda row: source_text(row.get("Status")) == "Unassigned",
        lambda row: source_text(row.get("Status")) == "Assigned to action tracker",
        lambda row: source_text(row.get("Incident Classification")) not in {None, "N/A"},
        lambda row: source_text(row.get("Property Damage")) == "Property Damage",
        lambda row: source_text(row.get("Site")) == "Kitwe",
        lambda row: source_text(row.get("Site")) == "Third Party Premises",
    )
    selected: list[str] = []
    for selector in selectors:
        row = next((candidate for candidate in rows if selector(candidate)), None)
        if row is None:
            continue
        external_id = source_text(row.get("ID"))
        if external_id and external_id not in selected:
            selected.append(external_id)
    return selected


def parity_samples(
    rows: list[dict[str, Any]],
    record_lookup: dict[str, SafetyImprovementObservation],
) -> list[dict[str, Any]]:
    source_lookup = {source_text(row["ID"]): row for row in rows}
    samples: list[dict[str, Any]] = []
    for external_id in representative_ids(rows):
        source_row = source_lookup[external_id]
        record = record_lookup.get(external_id)
        if record is None:
            samples.append(
                {
                    "external_reference_id": external_id,
                    "imported": False,
                    "differences": ["record not imported"],
                }
            )
            continue
        api_payload = SIORead.model_validate(record).model_dump(mode="json")
        differences = [
            field
            for field in YALELO_SIO_COLUMNS
            if not normalized_compare(
                source_row.get(field),
                destination_value(record, field),
                field=field,
            )
        ]
        samples.append(
            {
                "external_reference_id": external_id,
                "imported": True,
                "database_differences": differences,
                "api_contract_differences": [
                    key
                    for key in (
                        "external_reference_id",
                        "observation_date",
                        "department",
                        "source_type",
                        "description",
                        "incident_classification",
                        "status",
                        "observation_nature",
                        "responsible_department",
                        "responsible_hs_officer_name",
                        "urgency",
                        "category",
                        "responsible_person_name",
                        "source_created_at",
                        "property_damage",
                        "source_created_by",
                        "source_modified_by",
                        "source_path",
                        "legacy_metadata",
                    )
                    if key not in api_payload
                ],
            }
        )
    return samples


def search_filter_checks(
    db: Session,
    rows: list[dict[str, Any]],
    record_lookup: dict[str, SafetyImprovementObservation],
) -> dict[str, bool | None]:
    if not record_lookup:
        return {"not_run": None}
    records = list(record_lookup.values())
    sample = records[0]
    source_row = next(row for row in rows if source_text(row["ID"]) == sample.external_reference_id)

    def found(**kwargs) -> bool:
        return list_sios(db, limit=1, **kwargs)["total"] > 0

    def first_with(field: str):
        return next((record for record in records if getattr(record, field) is not None), None)

    responsible_department = first_with("responsible_department_id")
    officer = first_with("responsible_hs_officer_name")
    responsible_person = first_with("responsible_person_name")
    category = first_with("category")
    classification = first_with("incident_classification")
    urgency = first_with("urgency")
    observation_date = first_with("observation_date")
    source_created_at = first_with("source_created_at")

    return {
        "source_id_search": found(search=sample.external_reference_id),
        "site_filter": found(site_id=sample.site_id),
        "department_filter": found(department=sample.department),
        "responsible_department_filter": (
            found(responsible_department_id=responsible_department.responsible_department_id)
            if responsible_department
            else None
        ),
        "officer_search": (
            found(search=officer.responsible_hs_officer_name)
            if officer
            else None
        ),
        "responsible_person_search": (
            found(search=responsible_person.responsible_person_name)
            if responsible_person
            else None
        ),
        "source_filter": found(source_type=sample.source_type),
        "category_filter": found(category=category.category) if category else None,
        "classification_filter": (
            found(incident_classification=classification.incident_classification)
            if classification
            else None
        ),
        "nature_filter": found(observation_nature=sample.observation_nature),
        "urgency_filter": found(urgency=urgency.urgency) if urgency else None,
        "status_filter": found(status=sample.status),
        "observation_date_filter": (
            found(
                date_from=observation_date.observation_date,
                date_to=observation_date.observation_date,
            )
            if observation_date
            else None
        ),
        "source_created_date_filter": (
            found(
                source_created_from=source_created_at.source_created_at.date(),
                source_created_to=source_created_at.source_created_at.date(),
            )
            if source_created_at
            else None
        ),
        "description_search": found(search=source_text(source_row["Description of SIO"])[:40]),
    }


def export_checks(db: Session, expected_count: int) -> dict[str, Any]:
    output = export_sios_csv(db)
    rows = list(csv.DictReader(StringIO(output)))
    required = {
        "External Reference ID",
        "Observation Date",
        "Originating Department",
        "Responsible Department",
        "Responsible Person",
        "Source",
        "Category",
        "Nature",
        "Urgency",
        "Status",
        "Description",
        "Incident Classification",
        "Responsible H&S Officer",
        "Source Created At",
        "Property Damage",
        "Source Created By",
        "Source Modified By",
    }
    headers = set(rows[0].keys()) if rows else set()
    return {
        "row_count": len(rows),
        "expected_row_count": expected_count,
        "required_headers_present": sorted(required.intersection(headers)),
        "missing_required_headers": sorted(required.difference(headers)),
    }


def run_audit(workbook_path: Path) -> dict[str, Any]:
    rows = source_rows(workbook_path)
    engine = create_engine(
        "sqlite+pysqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SessionLocal = sessionmaker(
        bind=engine,
        class_=TenantSession,
        autoflush=False,
        autocommit=False,
    )
    Base.metadata.create_all(engine)
    db = SessionLocal()
    try:
        seeded = seed_tenant(
            db,
            organisation_id=1,
            organisation_name="Yalelo Pilot Audit",
            organisation_code="YALELO-AUDIT",
            sites=("Siavonga", "Lusaka", "Yalelo Stores", "Kitwe"),
            departments=MAPPED_DEPARTMENTS,
            users=(*EXACT_USER_NAMES, "Amos Lupupulala", "Amos Lupupulala"),
        )
        set_tenant_context(db, 1, platform_admin=True)
        before = counts(db)
        content = workbook_path.read_bytes()
        preview_before_sios = counts(db)["sios"]
        preview = preview_yalelo_sio_import(
            db,
            content=content,
            filename=workbook_path.name,
            actor_id=seeded["actor"].id,
        )
        preview_after_sios = counts(db)["sios"]
        preview_report = dict(preview.report)
        preview_report["operational_sios_before"] = preview_before_sios
        preview_report["operational_sios_after"] = preview_after_sios

        confirmed = confirm_yalelo_sio_import(
            db,
            preview,
            ImportConfirmRequest(create_sites=["Third Party Premises"]),
            actor_id=seeded["actor"].id,
        )
        after_first = counts(db)
        records = list(db.scalars(select(SafetyImprovementObservation)).unique().all())
        record_lookup = {record.external_reference_id or "": record for record in records}
        loss = data_loss_profile(rows, record_lookup)
        samples = parity_samples(rows, record_lookup)
        filters = search_filter_checks(db, rows, record_lookup)
        exported = export_checks(db, len(records))
        analytics = get_sio_analytics(db) if records else None

        second_preview = preview_yalelo_sio_import(
            db,
            content=content,
            filename=workbook_path.name,
            actor_id=seeded["actor"].id,
        )
        second_confirmed = confirm_yalelo_sio_import(
            db,
            second_preview,
            ImportConfirmRequest(),
            actor_id=seeded["actor"].id,
        )
        after_second = counts(db)

        first_sio_id = records[0].id if records else None
        first_external_id = records[0].external_reference_id if records else None
        org1_site_id = seeded["sites"][0].id
        org1_sio_count = after_second["sios"]
        org1_job_count = after_second["import_jobs"]

        tenant2 = seed_tenant(
            db,
            organisation_id=2,
            organisation_name="Isolation Control Tenant",
            organisation_code="CONTROL-AUDIT",
            sites=("Control Site",),
        )
        set_tenant_context(db, 2, platform_admin=True)
        tenant2_initial = counts(db)
        guessed_sio = db.get(SafetyImprovementObservation, first_sio_id) if first_sio_id else None
        guessed_job = db.get(DataImportJob, preview.id)
        same_external_id_created = False
        if first_external_id:
            control_sio = create_sio(
                db,
                SIOCreate(
                    external_reference_id=first_external_id,
                    source_system=YALELO_SOURCE_SYSTEM,
                    department="Control",
                    source_type="Audit",
                    description="Tenant-aware duplicate control record",
                    status="unassigned",
                    observation_nature="positive",
                    site_id=tenant2["sites"][0].id,
                ),
                actor_id=tenant2["actor"].id,
                is_import=True,
            )
            same_external_id_created = control_sio.id is not None
        tenant2_after = counts(db)
        tenant2_export_rows = len(list(csv.DictReader(StringIO(export_sios_csv(db)))))
        tenant2_dashboard_total = get_sio_analytics(db)["total_observations"]

        set_tenant_context(db, 1, platform_admin=True)
        org1_after_control = counts(db)
        cross_tenant_site_visible = db.get(Site, tenant2["sites"][0].id) is not None

        return {
            "database": {
                "kind": "disposable in-memory SQLite",
                "production_connected": False,
            },
            "source": {
                "path": str(workbook_path.resolve()),
                "rows": len(rows),
                "columns": len(YALELO_SIO_COLUMNS),
            },
            "counts_before": before,
            "preview": preview_report,
            "confirmation": dict(confirmed.report),
            "counts_after_first_import": after_first,
            "second_preview": dict(second_preview.report),
            "second_confirmation": dict(second_confirmed.report),
            "counts_after_second_import": after_second,
            "data_loss": loss,
            "total_failed_or_lost_populated_values": sum(
                item["failed_or_lost_count"] for item in loss
            ),
            "field_parity_samples": samples,
            "search_filter_checks": filters,
            "export_checks": exported,
            "analytics": analytics,
            "side_effects": {
                "actions_created": after_first["actions"] - before["actions"],
                "incidents_created": after_first["incidents"] - before["incidents"],
                "hazards_created": after_first["hazards"] - before["hazards"],
                "notifications_created": after_first["notifications"] - before["notifications"],
                "users_created_beyond_seed": after_first["users"] - before["users"],
                "departments_created_beyond_seed": after_first["departments"] - before["departments"],
                "sites_created_beyond_seed": after_first["sites"] - before["sites"],
            },
            "tenant_isolation": {
                "org1_sio_count_before_control": org1_sio_count,
                "org1_job_count_before_control": org1_job_count,
                "tenant2_initial_sio_count": tenant2_initial["sios"],
                "tenant2_initial_job_count": tenant2_initial["import_jobs"],
                "guessed_org1_sio_visible": guessed_sio is not None,
                "guessed_org1_import_job_visible": guessed_job is not None,
                "same_external_id_created_in_other_tenant": same_external_id_created,
                "tenant2_export_rows": tenant2_export_rows,
                "tenant2_dashboard_total": tenant2_dashboard_total,
                "org1_sio_count_after_control": org1_after_control["sios"],
                "org1_site_id": org1_site_id,
                "cross_tenant_site_visible_from_org1": cross_tenant_site_visible,
            },
        }
    finally:
        db.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Run a Yalelo import audit against a disposable in-memory database."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    payload = json.dumps(run_audit(args.workbook), indent=2, ensure_ascii=False, default=str)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


if __name__ == "__main__":
    main()
