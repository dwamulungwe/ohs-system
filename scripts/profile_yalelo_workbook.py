from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


EXPECTED_COLUMNS = (
    "ID",
    "Date",
    "Department",
    "Source of Observation",
    "Description of SIO",
    "Incident Classification",
    "Status",
    "Nature of Observation",
    "Department Responsible for Corrective Action",
    "Site",
    "Responsible H&S Officer",
    "Urgency",
    "SIO Category",
    "Person Responsible for Corrective Action",
    "Created",
    "Property Damage",
    "Created By",
    "Modified By",
    "MonthY",
    "Item Type",
    "Path",
)

EXACT_VALUE_COLUMNS = (
    "Site",
    "Status",
    "Nature of Observation",
    "Urgency",
    "Source of Observation",
    "Incident Classification",
    "Department Responsible for Corrective Action",
    "Responsible H&S Officer",
    "Person Responsible for Corrective Action",
    "Property Damage",
    "MonthY",
    "Item Type",
    "Path",
)

COUNT_COLUMNS = (
    "SIO Category",
    "Department",
)

STRING_DATE_FORMATS = (
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y/%m/%d",
    "%d-%b-%Y",
    "%d-%m-%Y",
)

MONTH_PATTERNS = (
    re.compile(r"^(?:19|20)\d{2}[-/]?(?:0?[1-9]|1[0-2])$"),
    re.compile(r"^(?:0?[1-9]|1[0-2])[-/](?:19|20)\d{2}$"),
    re.compile(
        r"^(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|"
        r"dec(?:ember)?)[- /](?:19|20)\d{2}$",
        re.IGNORECASE,
    ),
)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return str(value)
    return value


def canonical_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def parse_date(value: Any, *, epoch: datetime) -> date | datetime | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = from_excel(value, epoch=epoch)
        if isinstance(parsed, time):
            raise ValueError("Excel serial resolves to a time-only value")
        return parsed
    if isinstance(value, str):
        text_value = value.strip()
        try:
            return datetime.fromisoformat(text_value.replace("Z", "+00:00"))
        except ValueError:
            for pattern in STRING_DATE_FORMATS:
                try:
                    return datetime.strptime(text_value, pattern)
                except ValueError:
                    continue
    raise ValueError(f"Unsupported date value: {value!r}")


def value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__


def month_quality(value: Any) -> str:
    if is_blank(value):
        return "blank"
    if isinstance(value, (datetime, date)):
        return "artifact" if value.year == 1899 else "valid_native_date"
    text_value = canonical_text(value)
    lowered = text_value.lower()
    if "1899" in lowered or lowered in {"0", "0.0"}:
        return "artifact"
    if any(pattern.fullmatch(text_value) for pattern in MONTH_PATTERNS):
        return "valid_string"
    return "inconsistent"


def profile_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_profiles: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                sheet_profiles.append(
                    {
                        "name": worksheet.title,
                        "row_count": 0,
                        "column_count": 0,
                        "headers": [],
                    }
                )
                continue

            headers = [canonical_text(value) if not is_blank(value) else None for value in header_row]
            header_counter = Counter(header for header in headers if header is not None)
            duplicate_columns = sorted(
                header for header, count in header_counter.items() if count > 1
            )
            missing_columns = [column for column in EXPECTED_COLUMNS if column not in headers]
            additional_columns = [
                header for header in headers if header is not None and header not in EXPECTED_COLUMNS
            ]

            raw_rows: list[tuple[Any, ...]] = []
            completely_blank_rows = 0
            malformed_width_rows: list[dict[str, int]] = []
            for row_number, values in enumerate(rows, start=2):
                if all(is_blank(value) for value in values):
                    completely_blank_rows += 1
                    continue
                if len(values) != len(headers):
                    malformed_width_rows.append(
                        {"row_number": row_number, "cell_count": len(values)}
                    )
                raw_rows.append(values)

            row_count = len(raw_rows)
            column_profiles: dict[str, dict[str, Any]] = {}
            column_values: dict[str, list[Any]] = {column: [] for column in EXPECTED_COLUMNS}
            for values in raw_rows:
                for index, header in enumerate(headers):
                    if header in column_values:
                        column_values[header].append(values[index] if index < len(values) else None)

            for column in EXPECTED_COLUMNS:
                values = column_values[column]
                populated = [value for value in values if not is_blank(value)]
                distinct = Counter(canonical_text(value) for value in populated)
                type_counts = Counter(value_type(value) for value in populated)
                malformed: list[dict[str, Any]] = []
                if column in {"Date", "Created"}:
                    parsed_values: list[date | datetime] = []
                    for index, value in enumerate(values, start=2):
                        if is_blank(value):
                            continue
                        try:
                            parsed = parse_date(value, epoch=workbook.epoch)
                            if parsed is not None:
                                parsed_values.append(parsed)
                            if parsed is not None and parsed.year == 1899:
                                malformed.append(
                                    {
                                        "row_number": index,
                                        "value": json_value(value),
                                        "reason": "Excel zero-date artifact",
                                    }
                                )
                        except (OverflowError, ValueError) as exc:
                            malformed.append(
                                {
                                    "row_number": index,
                                    "value": json_value(value),
                                    "reason": str(exc),
                                }
                            )
                    if parsed_values:
                        parsed_datetimes = [
                            value
                            if isinstance(value, datetime)
                            else datetime.combine(value, time.min)
                            for value in parsed_values
                        ]
                        column_date_summary = {
                            "minimum": min(parsed_datetimes).isoformat(),
                            "maximum": max(parsed_datetimes).isoformat(),
                            "timezone_aware_count": sum(
                                value.tzinfo is not None for value in parsed_datetimes
                            ),
                            "timezone_naive_count": sum(
                                value.tzinfo is None for value in parsed_datetimes
                            ),
                        }
                    else:
                        column_date_summary = None
                elif column == "MonthY":
                    for index, value in enumerate(values, start=2):
                        quality = month_quality(value)
                        if quality in {"artifact", "inconsistent"}:
                            malformed.append(
                                {
                                    "row_number": index,
                                    "value": json_value(value),
                                    "reason": quality,
                                }
                            )
                elif column == "ID":
                    for index, value in enumerate(values, start=2):
                        if is_blank(value):
                            continue
                        text_value = canonical_text(value)
                        if not text_value or any(char in text_value for char in "\r\n\t"):
                            malformed.append(
                                {
                                    "row_number": index,
                                    "value": json_value(value),
                                    "reason": "blank/control-character identifier",
                                }
                            )

                profile: dict[str, Any] = {
                    "populated_rows": len(populated),
                    "blank_rows": row_count - len(populated),
                    "null_percentage": round(
                        ((row_count - len(populated)) / row_count * 100) if row_count else 0,
                        3,
                    ),
                    "distinct_count": len(distinct),
                    "type_counts": dict(sorted(type_counts.items())),
                    "malformed_count": len(malformed),
                    "malformed_values": malformed[:50],
                }
                if column in {"Date", "Created"}:
                    profile["date_summary"] = column_date_summary
                if populated and all(isinstance(value, str) for value in populated):
                    lengths = [len(value) for value in populated]
                    profile["max_length"] = max(lengths)
                    profile["leading_or_trailing_whitespace_count"] = sum(
                        value != value.strip() for value in populated
                    )
                    profile["line_break_count"] = sum(
                        "\n" in value or "\r" in value for value in populated
                    )
                    profile["unicode_count"] = sum(
                        any(ord(character) > 127 for character in value) for value in populated
                    )
                column_profiles[column] = profile

            source_ids = [
                canonical_text(value)
                for value in column_values["ID"]
                if not is_blank(value)
            ]
            exact_id_counts = Counter(source_ids)
            normalized_id_counts = Counter(value.casefold() for value in source_ids)

            exact_values = {
                column: [
                    {"value": value, "count": count}
                    for value, count in sorted(
                        Counter(
                            canonical_text(value)
                            for value in column_values[column]
                            if not is_blank(value)
                        ).items(),
                        key=lambda item: (item[0].casefold(), item[0]),
                    )
                ]
                for column in EXACT_VALUE_COLUMNS
            }
            value_counts = {
                column: [
                    {"value": value, "count": count}
                    for value, count in Counter(
                        canonical_text(value)
                        for value in column_values[column]
                        if not is_blank(value)
                    ).most_common()
                ]
                for column in COUNT_COLUMNS
            }

            description_values = [
                value
                for value in column_values["Description of SIO"]
                if isinstance(value, str) and value
            ]
            property_damage_values = [
                value for value in column_values["Property Damage"] if not is_blank(value)
            ]
            month_quality_counts = Counter(
                month_quality(value) for value in column_values["MonthY"]
            )
            completely_empty_columns = [
                column
                for column in EXPECTED_COLUMNS
                if not any(not is_blank(value) for value in column_values[column])
            ]

            sheet_profiles.append(
                {
                    "name": worksheet.title,
                    "row_count": row_count,
                    "column_count": len(headers),
                    "headers": headers,
                    "contract": {
                        "exact_match": headers == list(EXPECTED_COLUMNS),
                        "missing_columns": missing_columns,
                        "additional_columns": additional_columns,
                        "duplicate_columns": duplicate_columns,
                        "completely_empty_columns": completely_empty_columns,
                    },
                    "duplicate_source_ids": [
                        {"value": value, "count": count}
                        for value, count in exact_id_counts.items()
                        if count > 1
                    ],
                    "case_insensitive_duplicate_source_ids": [
                        {"value": value, "count": count}
                        for value, count in normalized_id_counts.items()
                        if count > 1
                    ],
                    "blank_ids": column_profiles["ID"]["blank_rows"],
                    "completely_blank_rows": completely_blank_rows,
                    "malformed_width_rows": malformed_width_rows,
                    "column_profiles": column_profiles,
                    "exact_values": exact_values,
                    "value_counts": value_counts,
                    "description_quality": {
                        "max_length": max((len(value) for value in description_values), default=0),
                        "line_break_count": sum(
                            "\n" in value or "\r" in value for value in description_values
                        ),
                        "unicode_count": sum(
                            any(ord(character) > 127 for character in value)
                            for value in description_values
                        ),
                        "leading_or_trailing_whitespace_count": sum(
                            value != value.strip() for value in description_values
                        ),
                    },
                    "property_damage_semantics": {
                        "type_counts": dict(
                            sorted(Counter(value_type(value) for value in property_damage_values).items())
                        ),
                        "distinct_count": len(
                            {canonical_text(value) for value in property_damage_values}
                        ),
                    },
                    "monthy_quality": dict(sorted(month_quality_counts.items())),
                }
            )

        return {
            "workbook": {
                "path": str(path.resolve()),
                "size_bytes": path.stat().st_size,
                "worksheet_names": workbook.sheetnames,
                "excel_epoch": workbook.epoch.isoformat(),
            },
            "sheets": sheet_profiles,
        }
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Profile a Yalelo SIO workbook without modifying it.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    payload = json.dumps(profile_workbook(args.workbook), indent=2, ensure_ascii=False)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


if __name__ == "__main__":
    main()
