from datetime import date, datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.corrective_action import (
    CorrectiveAction,
    CorrectiveActionPriority,
    CorrectiveActionStatus,
)
from app.models.incident import Incident, IncidentSeverity
from app.models.organisation import OrganisationFeature
from app.models.reporting import KPISnapshot, ReportingPeriod, ReportingPeriodStatus


def _create_period(client, *, name="April 2026 Management Report", start="2026-04-01", end="2026-04-30"):
    response = client.post(
        "/api/v1/reporting/periods",
        json={"name": name, "period_type": "monthly", "start_date": start, "end_date": end},
    )
    assert response.status_code == 201, response.text
    return response.json()


def _seed_action_and_incident(db_session):
    action = CorrectiveAction(
        action_reference="ACT-2026-000001",
        site_id=1,
        title="Close critical guarding gap",
        description="Install and verify the guard",
        priority=CorrectiveActionPriority.critical,
        lifecycle_status=CorrectiveActionStatus.in_progress,
        original_due_date=date(2026, 4, 10),
        current_due_date=date(2026, 4, 10),
        created_by_user_id=1,
    )
    incident = Incident(
        site_id=1,
        title="Recordable hand injury",
        description="Recordable case",
        severity=IncidentSeverity.medium,
        occurred_at=datetime(2026, 4, 12, tzinfo=timezone.utc),
        is_recordable=True,
        reported_by_id=1,
    )
    db_session.add_all([action, incident])
    db_session.commit()
    return action, incident


def test_period_lifecycle_snapshot_immutability_and_restatement(client, db_session):
    action, _ = _seed_action_and_incident(db_session)
    exposure = client.post(
        "/api/v1/reporting/workforce-exposure",
        json={
            "period_start": "2026-04-01",
            "period_end": "2026-04-30",
            "employee_hours_worked": 1000,
            "contractor_hours_worked": 500,
        },
    )
    assert exposure.status_code == 201, exposure.text
    period = _create_period(client)

    generated = client.post(f"/api/v1/reporting/periods/{period['id']}/snapshots")
    assert generated.status_code == 200, generated.text
    scorecard = client.get(f"/api/v1/reporting/periods/{period['id']}/scorecard").json()
    values = {row["kpi_key"]: row for row in scorecard["rows"]}
    assert values["action_open"]["actual"] == 1
    assert values["action_overdue"]["actual"] == 1
    assert values["trir"]["actual"] == pytest.approx(133.33)
    assert values["trir"]["denominator"] == 1500
    drilldown = client.get(
        f"/api/v1/reporting/periods/{period['id']}/kpis/action_overdue/drilldown"
    )
    assert drilldown.status_code == 200
    assert drilldown.json()["items"][0]["reference"] == action.action_reference

    for command in ("submit", "review", "approve", "lock"):
        response = client.post(f"/api/v1/reporting/periods/{period['id']}/{command}")
        assert response.status_code == 200, response.text
    locked = client.get(f"/api/v1/reporting/periods/{period['id']}").json()
    assert locked["status"] == "locked"
    assert locked["report_reference"] == "HSE-MR-2026-04-V1"
    assert client.post(f"/api/v1/reporting/periods/{period['id']}/snapshots").status_code == 423

    original_value = values["action_open"]["actual"]
    action.lifecycle_status = CorrectiveActionStatus.closed
    action.closed_at = datetime(2026, 5, 2, tzinfo=timezone.utc)
    db_session.commit()
    assert {
        row["kpi_key"]: row["actual"]
        for row in client.get(f"/api/v1/reporting/periods/{period['id']}/scorecard").json()["rows"]
    }["action_open"] == original_value

    restated = client.post(
        f"/api/v1/reporting/periods/{period['id']}/reopen",
        json={"reason": "Correct late workforce hours"},
    )
    assert restated.status_code == 200, restated.text
    assert restated.json()["id"] != period["id"]
    assert restated.json()["report_version"] == 2
    assert restated.json()["supersedes_period_id"] == period["id"]
    assert client.get(f"/api/v1/reporting/periods/{period['id']}").json()["status"] == "locked"


def test_trir_and_ltifr_never_fabricate_missing_denominators(client, db_session):
    _seed_action_and_incident(db_session)
    response = client.post(
        "/api/v1/reporting/workforce-exposure",
        json={
            "period_start": "2026-04-01",
            "period_end": "2026-04-30",
            "employee_hours_worked": 1000,
        },
    )
    assert response.status_code == 201
    period = _create_period(client)
    assert client.post(f"/api/v1/reporting/periods/{period['id']}/snapshots").status_code == 200
    values = {
        row["kpi_key"]: row
        for row in client.get(f"/api/v1/reporting/periods/{period['id']}/scorecard").json()["rows"]
    }
    assert values["trir"]["actual"] is None
    assert values["trir"]["status"] == "insufficient_data"
    assert "hours" in values["trir"]["explanation"]["insufficient_data_reason"].lower()
    assert values["ltifr"]["actual"] is None


def test_effective_target_is_copied_into_snapshot(client):
    definitions = client.get("/api/v1/reporting/kpi-definitions", params={"key": "action_overdue"}).json()
    definition = definitions[0]
    first = client.post(
        "/api/v1/reporting/kpi-targets",
        json={
            "kpi_definition_id": definition["id"],
            "target_value": 0,
            "warning_threshold": 1,
            "critical_threshold": 5,
            "effective_from": "2026-01-01",
        },
    )
    assert first.status_code == 201, first.text
    second = client.post(
        "/api/v1/reporting/kpi-targets",
        json={
            "kpi_definition_id": definition["id"],
            "target_value": 2,
            "warning_threshold": 3,
            "critical_threshold": 8,
            "effective_from": "2026-05-01",
        },
    )
    assert second.status_code == 201, second.text
    assert second.json()["version"] == 2

    period = _create_period(client)
    assert client.post(f"/api/v1/reporting/periods/{period['id']}/snapshots").status_code == 200
    overdue = next(
        row for row in client.get(f"/api/v1/reporting/periods/{period['id']}/scorecard").json()["rows"]
        if row["kpi_key"] == "action_overdue"
    )
    assert overdue["target"] == 0
    assert overdue["explanation"]["target"]["target_version"] == 1


def test_feature_disabled_sections_and_reporting_rbac(client, db_session, create_user_for_role, act_as):
    hazards_feature = db_session.scalar(select(OrganisationFeature).where(OrganisationFeature.key == "hazards"))
    hazards_feature.is_enabled = False
    db_session.commit()
    period = _create_period(client)
    sections = client.get(f"/api/v1/reporting/periods/{period['id']}/sections").json()
    assert next(item for item in sections if item["section_key"] == "risk_hazards")["is_enabled"] is False

    safety_officer = create_user_for_role("safety_officer", assigned_site_id=1)
    act_as(safety_officer)
    assert client.get("/api/v1/reporting/periods").status_code == 200
    assert client.post(
        "/api/v1/reporting/periods",
        json={"name": "Denied", "period_type": "monthly", "start_date": "2026-05-01", "end_date": "2026-05-31"},
    ).status_code == 403


def test_report_exports_contain_expected_pdf_and_excel_structure(client, db_session):
    _seed_action_and_incident(db_session)
    period = _create_period(client)
    assert client.post(f"/api/v1/reporting/periods/{period['id']}/snapshots").status_code == 200
    pdf = client.get(f"/api/v1/reporting/periods/{period['id']}/exports/pdf")
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"] == "application/pdf"
    assert pdf.content.startswith(b"%PDF-1.4")

    excel = client.get(f"/api/v1/reporting/periods/{period['id']}/exports/excel")
    assert excel.status_code == 200, excel.text
    workbook = load_workbook(BytesIO(excel.content), read_only=True)
    assert {
        "Scorecard", "SIO", "Incidents", "Actions", "Hazards", "Inspections",
        "Audits", "Training", "Compliance", "Forward View", "Management Action Plan",
    }.issubset(workbook.sheetnames)


def test_locked_snapshot_direct_mutation_is_rejected(client, db_session):
    period = _create_period(client)
    assert client.post(f"/api/v1/reporting/periods/{period['id']}/snapshots").status_code == 200
    for command in ("submit", "review", "approve", "lock"):
        assert client.post(f"/api/v1/reporting/periods/{period['id']}/{command}").status_code == 200
    snapshot = db_session.scalar(select(KPISnapshot).where(KPISnapshot.reporting_period_id == period["id"]))
    snapshot.value = 999
    with pytest.raises(RuntimeError, match="immutable"):
        db_session.commit()
    db_session.rollback()
    persisted = db_session.scalar(select(KPISnapshot).where(KPISnapshot.id == snapshot.id))
    assert persisted.value != 999


def test_period_validation_and_duplicate_policy(client):
    invalid = client.post(
        "/api/v1/reporting/periods",
        json={"name": "Invalid", "period_type": "monthly", "start_date": "2026-05-31", "end_date": "2026-05-01"},
    )
    assert invalid.status_code == 422
    period = _create_period(client)
    duplicate = client.post(
        "/api/v1/reporting/periods",
        json={"name": period["name"], "period_type": period["period_type"], "start_date": period["start_date"], "end_date": period["end_date"]},
    )
    assert duplicate.status_code == 409
