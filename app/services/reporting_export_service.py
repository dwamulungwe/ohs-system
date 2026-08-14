from __future__ import annotations

import hashlib
from io import BytesIO
from typing import Iterable, Optional

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.organisation import Organisation
from app.models.reporting import ReportExport, ReportSection, ReportingPeriod, ReportingPeriodHistory
from app.services.audit_service import write_audit_log
from app.services.reporting_insights import get_forward_view
from app.services.reporting_service import get_scorecard, list_report_sections


SHEET_CATEGORIES = {
    "SIO": ("sio_",),
    "Incidents": (
        "total_incidents", "near_miss_count", "first_aid_count", "medical_treatment_count",
        "restricted_work_count", "lost_time_injury_count", "occupational_illness_count",
        "fatality_count", "property_damage_count", "environmental_incident_count",
        "high_critical_incidents", "open_investigations", "overdue_investigations",
        "average_investigation_closure_days", "days_since_last_lti", "trir", "ltifr",
    ),
    "Actions": ("action_",),
    "Hazards": (
        "open_hazards", "critical_hazards", "high_risk_hazards", "uncontrolled_hazards",
        "overdue_controls", "hazards_due_review", "new_hazards", "hazards_closed",
        "residual_high_risk_hazards",
    ),
    "Inspections": ("inspection", "critical_inspection", "repeat_inspection"),
    "Audits": ("audit", "audits_", "major_findings", "minor_findings", "open_audit", "overdue_audit", "repeat_audit"),
    "Training": ("training", "competency_gaps"),
    "Compliance": ("permit", "permits", "active_permits", "expired_permits", "compliance"),
}


def _organisation(db: Session, period: ReportingPeriod) -> Organisation:
    organisation = db.get(Organisation, period.organisation_id)
    if organisation is None:
        raise RuntimeError("Reporting organisation was not found")
    return organisation


def _matches(key: str, selectors: tuple[str, ...]) -> bool:
    return any(key == selector or key.startswith(selector) for selector in selectors)


def _display(value, unit: Optional[str] = None) -> str:
    if value is None:
        return "Unavailable"
    if unit == "percent":
        return f"{value:.2f}%"
    if unit == "days":
        return f"{value:.2f} days"
    return f"{value:.2f}" if isinstance(value, float) and not value.is_integer() else str(int(value) if isinstance(value, float) else value)


def build_report_xlsx(db: Session, period: ReportingPeriod) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    organisation = _organisation(db, period)
    scorecard = get_scorecard(db, period)
    workbook = Workbook()
    scorecard_sheet = workbook.active
    scorecard_sheet.title = "Scorecard"
    header_fill = PatternFill("solid", fgColor="166534")
    header_font = Font(color="FFFFFF", bold=True)

    info = workbook.create_sheet("Report Info", 0)
    info_rows = [
        ("Organisation", organisation.name),
        ("Reporting Period", period.name),
        ("Period", f"{period.start_date.isoformat()} to {period.end_date.isoformat()}"),
        ("Report Reference", period.report_reference or f"Draft V{period.report_version}"),
        ("Report Version", period.report_version),
        ("Status", period.status.value),
        ("Prepared By", period.prepared_by_name or ""),
        ("Reviewed By", period.reviewed_by_name or ""),
        ("Approved By", period.approved_by_name or ""),
        ("Approval Date", period.approved_at.isoformat() if period.approved_at else ""),
        ("Locked Date", period.locked_at.isoformat() if period.locked_at else ""),
    ]
    for row in info_rows:
        info.append(row)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 54
    for cell in info[1]:
        cell.font = Font(bold=True)

    headers = ["KPI", "Target", "Actual", "Previous Period", "YTD", "Status", "Numerator", "Denominator", "Unit"]
    scorecard_sheet.append(headers)
    for row in scorecard["rows"]:
        scorecard_sheet.append([
            row["kpi_name"], row["target"], row["actual"], row["previous_period"],
            row["ytd"], row["status"].value if hasattr(row["status"], "value") else row["status"],
            row["numerator"], row["denominator"], row["unit"],
        ])
    for cell in scorecard_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    scorecard_sheet.freeze_panes = "A2"
    scorecard_sheet.auto_filter.ref = scorecard_sheet.dimensions

    for title, selectors in SHEET_CATEGORIES.items():
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in scorecard["rows"]:
            if _matches(row["kpi_key"], selectors):
                sheet.append([
                    row["kpi_name"], row["target"], row["actual"], row["previous_period"],
                    row["ytd"], row["status"].value if hasattr(row["status"], "value") else row["status"],
                    row["numerator"], row["denominator"], row["unit"],
                ])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    forward = workbook.create_sheet("Forward View")
    forward.append(["Source", "Title", "Due Date", "Days Until Due", "Window", "Site", "Department"])
    for item in get_forward_view(db, as_of=period.end_date, window_days=90):
        forward.append([
            item["source_type"], item["title"], item["obligation_date"], item["days_until_due"],
            item["window_days"], item["site_id"], item["department_id"],
        ])
    for cell in forward[1]:
        cell.fill = header_fill
        cell.font = header_font

    summary = workbook.create_sheet("Executive Summary")
    section = next((item for item in list_report_sections(db, period) if item.section_key == "executive_summary"), None)
    for key, value in (section.content if section else {}).items():
        summary.append([key.replace("_", " ").title(), value])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 100
    for row in summary.iter_rows():
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    action_plan = workbook.create_sheet("Management Action Plan")
    action_plan.append(["Action Reference", "Priority", "Issue", "Management Comment", "Owner", "Due Date", "Status"])
    for item in period.management_actions:
        action = item.linked_action
        action_plan.append([
            action.action_reference, item.priority, item.issue_summary, item.management_comment,
            action.owner_name, action.current_due_date, action.lifecycle_status.value,
        ])
    for cell in action_plan[1]:
        cell.fill = header_fill
        cell.font = header_font

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[letter].width = width

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _pdf_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _minimal_pdf(lines: Iterable[str]) -> bytes:
    prepared = [str(line).encode("latin-1", "replace").decode("latin-1") for line in lines]
    pages = [prepared[index:index + 48] for index in range(0, len(prepared), 48)] or [[]]
    font_object_id = 3 + len(pages) * 2
    objects: list[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    kids = " ".join(f"{3 + index * 2} 0 R" for index in range(len(pages)))
    objects.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>".encode("ascii"))
    for index, page_lines in enumerate(pages):
        page_id = 3 + index * 2
        content_id = page_id + 1
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 {font_object_id} 0 R >> >> /Contents {content_id} 0 R >>".encode("ascii")
        )
        commands = ["BT", "/F1 9 Tf", "45 805 Td", "13 TL"]
        for line in page_lines:
            commands.extend([f"({_pdf_escape(line[:115])}) Tj", "T*"])
        commands.append("ET")
        content = "\n".join(commands).encode("latin-1")
        objects.append(f"<< /Length {len(content)} >>\nstream\n".encode("ascii") + content + b"\nendstream")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for object_id, body in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{object_id} 0 obj\n".encode("ascii"))
        output.extend(body)
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    return bytes(output)


def build_report_pdf(db: Session, period: ReportingPeriod) -> bytes:
    organisation = _organisation(db, period)
    scorecard = get_scorecard(db, period)
    lines = [
        organisation.name,
        "EXECUTIVE HSE MANAGEMENT REPORT",
        f"Reporting period: {period.name} ({period.start_date.isoformat()} to {period.end_date.isoformat()})",
        f"Report reference: {period.report_reference or f'DRAFT-V{period.report_version}'}",
        f"Status: {period.status.value}",
        "",
        "APPROVAL METADATA",
        f"Prepared By: {period.prepared_by_name or 'Not recorded'}",
        f"Reviewed By: {period.reviewed_by_name or 'Not recorded'}",
        f"Approved By: {period.approved_by_name or 'Not recorded'}",
        f"Approval Date: {period.approved_at.isoformat() if period.approved_at else 'Not approved'}",
        "",
        "EXECUTIVE SUMMARY",
    ]
    section = next((item for item in list_report_sections(db, period) if item.section_key == "executive_summary"), None)
    for key, value in (section.content if section else {}).items():
        lines.append(f"{key.replace('_', ' ').title()}: {value or '-'}")
    lines.extend(["", "KPI SCORECARD", "KPI | Target | Actual | Previous | YTD | Status"])
    for row in scorecard["rows"]:
        status = row["status"].value if hasattr(row["status"], "value") else row["status"]
        lines.append(
            f"{row['kpi_name']} | {_display(row['target'], row['unit'])} | {_display(row['actual'], row['unit'])} | "
            f"{_display(row['previous_period'], row['unit'])} | {_display(row['ytd'], row['unit'])} | {status}"
        )
    lines.extend(["", "MANAGEMENT ACTION PLAN"])
    for item in period.management_actions:
        action = item.linked_action
        lines.append(
            f"{action.action_reference}: {item.issue_summary} | Priority {item.priority} | "
            f"Owner {action.owner_name or '-'} | Due {_display(action.current_due_date)} | {action.lifecycle_status.value}"
        )
    return _minimal_pdf(lines)


def record_report_export(
    db: Session,
    period: ReportingPeriod,
    *,
    actor_id: int,
    export_format: str,
    file_name: str,
    content: bytes,
) -> ReportExport:
    record = ReportExport(
        reporting_period_id=period.id,
        export_format=export_format,
        file_name=file_name,
        checksum_sha256=hashlib.sha256(content).hexdigest(),
        report_version=period.report_version,
        generated_by_user_id=actor_id,
    )
    db.add(record)
    db.add(
        ReportingPeriodHistory(
            reporting_period_id=period.id,
            actor_user_id=actor_id,
            event_type="exported",
            from_status=period.status.value,
            to_status=period.status.value,
            event_metadata={
                "format": export_format,
                "file_name": file_name,
                "checksum_sha256": record.checksum_sha256,
                "report_version": period.report_version,
            },
        )
    )
    db.commit()
    db.refresh(record)
    write_audit_log(
        db,
        actor_id=actor_id,
        action="report.export",
        resource_type="reporting_period",
        resource_id=period.id,
        details={"format": export_format, "file_name": file_name, "checksum_sha256": record.checksum_sha256},
    )
    return record
