from __future__ import annotations

import re
from collections import Counter
from datetime import date, datetime, time, timezone
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.data_import import (
    DataImportJob,
    DataImportRow,
    ImportJobStatus,
    ImportRowStatus,
)
from app.models.department import Department
from app.models.site import Site
from app.models.sio import SIOObservationNature, SIOStatus, SIOUrgency, SafetyImprovementObservation
from app.models.user import User
from app.schemas.data_import import ImportConfirmRequest
from app.schemas.sio import SIOCreate
from app.services.audit_service import write_audit_log
from app.services.query_utils import paginate
from app.services.sio_service import SIODuplicateError, create_sio


YALELO_SIO_IMPORTER = "yalelo_sio"
YALELO_SOURCE_SYSTEM = "yalelo_sharepoint"
YALELO_SIO_COLUMNS = (
    "ID",
    "Date",
    "Department",
    "Source of Observation",
    "Description of SIO",
    "Incident Classification",
    "Status",
    "Nature of Observation",
    "Department Responsible for Corrective Action",
    "Site",
    "Responsible H&S Officer",
    "Urgency",
    "SIO Category",
    "Person Responsible for Corrective Action",
    "Created",
    "Property Damage",
    "Created By",
    "Modified By",
    "MonthY",
    "Item Type",
    "Path",
)
YALELO_REQUIRED_VALUES = (
    "ID",
    "Department",
    "Source of Observation",
    "Description of SIO",
    "Nature of Observation",
    "Site",
)
YALELO_SUPPORTED_SITES = (
    "Siavonga",
    "Lusaka",
    "Yalelo Stores",
    "Kitwe",
    "Third Party Premises",
)


class DataImportError(Exception):
    pass


class ImportJobNotFoundError(DataImportError):
    pass


class ImportValidationError(DataImportError):
    pass


class ImportStateError(DataImportError):
    pass


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _key(value: Any) -> str:
    text = str(value or "").strip().lower().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


STATUS_MAP = {
    "unassigned": SIOStatus.unassigned,
    "assigned_to_responsible_person": SIOStatus.assigned_to_responsible_person,
    "assigned_to_action_tracker": SIOStatus.assigned_to_action_tracker,
    "complete": SIOStatus.complete,
    "completed": SIOStatus.complete,
    "no_action_required": SIOStatus.no_action_required,
    "no_action": SIOStatus.no_action_required,
    "open": SIOStatus.open,
}
NATURE_MAP = {
    "positive": SIOObservationNature.positive,
    "positive_observation": SIOObservationNature.positive,
    "positive_safety_observation": SIOObservationNature.positive,
    "negative": SIOObservationNature.negative,
    "negative_observation": SIOObservationNature.negative,
    "negative_safety_observation": SIOObservationNature.negative,
}
URGENCY_MAP = {
    "low": SIOUrgency.low,
    "medium": SIOUrgency.medium,
    "normal": SIOUrgency.medium,
    "high": SIOUrgency.high,
    "urgent": SIOUrgency.urgent,
    "critical": SIOUrgency.urgent,
    "n_a": SIOUrgency.not_applicable,
    "na": SIOUrgency.not_applicable,
    "not_applicable": SIOUrgency.not_applicable,
}

YALELO_KNOWN_INCIDENT_CLASSIFICATIONS = {
    "fatality",
    "fire",
    "first_aid_injury",
    "lti",
    "medical_treatment_injury",
    "n_a",
    "near_miss",
    "occupational_illness",
    "other",
    "product_loss",
    "property_damage",
    "undesired_circumstance",
}


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _preserved_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    return text if text.strip() else None


def parse_excel_date(value: Any, *, epoch, as_datetime: bool) -> Optional[date | datetime]:
    if value is None or value == "":
        return None
    parsed: date | datetime
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = from_excel(value, epoch=epoch)
    elif isinstance(value, str):
        text_value = value.strip()
        if not text_value:
            return None
        try:
            parsed = datetime.fromisoformat(text_value.replace("Z", "+00:00"))
        except ValueError:
            parsed = None
            for pattern in ("%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%b-%Y", "%d-%m-%Y"):
                try:
                    parsed = datetime.strptime(text_value, pattern)
                    break
                except ValueError:
                    continue
            if parsed is None:
                raise ValueError(f"Unsupported Excel date value: {value}")
    else:
        raise ValueError(f"Unsupported Excel date value: {value}")

    if isinstance(parsed, time):
        raise ValueError(f"Excel date value resolves to time only: {value}")

    if as_datetime:
        result = parsed if isinstance(parsed, datetime) else datetime.combine(parsed, time.min)
        return result if result.tzinfo is not None else result.replace(tzinfo=timezone.utc)
    return parsed.date() if isinstance(parsed, datetime) else parsed


def _enum_value(value: Any, mapping: dict, field: str, *, optional: bool = False):
    if value is None or not str(value).strip():
        if optional:
            return None
        raise ValueError(f"{field} is required")
    normalized = mapping.get(_key(value))
    if normalized is None:
        raise ValueError(f"Unrecognized {field}: {value}")
    return normalized


def _exact_user_match(db: Session, name: Optional[str]) -> tuple[Optional[int], str]:
    if not name:
        return None, "blank"
    matches = list(
        db.scalars(
            select(User).where(func.lower(User.full_name) == name.strip().lower())
        ).all()
    )
    if len(matches) == 1:
        return matches[0].id, "resolved"
    return None, "ambiguous" if len(matches) > 1 else "unresolved"


def _exact_department_match(db: Session, name: Optional[str]) -> tuple[Optional[int], str]:
    if not name:
        return None, "blank"
    normalized = name.strip().lower()
    matches = list(
        db.scalars(
            select(Department).where(
                (func.lower(Department.name) == normalized)
                | (func.lower(Department.code) == normalized)
            )
        ).all()
    )
    if len(matches) == 1:
        return matches[0].id, "resolved"
    return None, "ambiguous" if len(matches) > 1 else "unresolved"


def _mapping_message(field: str, value: Optional[str], match: str) -> Optional[dict]:
    if not value or match in {"blank", "resolved"}:
        return None
    if match == "ambiguous":
        message = "Multiple exact tenant matches exist; source text was preserved without mapping"
    else:
        message = "No exact tenant match exists; source text was preserved without mapping"
    return {
        "field": field,
        "level": "warning",
        "code": f"{match}_mapping",
        "source_value": value,
        "message": message,
    }


def _site_lookup(db: Session) -> tuple[dict[str, int], set[str]]:
    by_name: dict[str, list[int]] = {}
    for site in db.scalars(select(Site)).all():
        by_name.setdefault(site.name.strip().lower(), []).append(site.id)
    resolved = {name: ids[0] for name, ids in by_name.items() if len(ids) == 1}
    ambiguous = {name for name, ids in by_name.items() if len(ids) > 1}
    return resolved, ambiguous


def _normalize_yalelo_row(db: Session, raw: dict[str, Any], *, epoch) -> tuple[dict, list[dict]]:
    messages: list[dict] = []
    for column in YALELO_REQUIRED_VALUES:
        if _text(raw.get(column)) is None:
            messages.append({"field": column, "level": "error", "message": f"{column} is required"})

    observation_date = None
    source_created_at = None
    status = None
    nature = None
    urgency = None
    for source_field, target_field, as_datetime in (
        ("Date", "observation_date", False),
        ("Created", "source_created_at", True),
    ):
        try:
            parsed = parse_excel_date(raw.get(source_field), epoch=epoch, as_datetime=as_datetime)
            if target_field == "observation_date":
                observation_date = parsed
            else:
                source_created_at = parsed
        except (ValueError, OverflowError) as exc:
            messages.append({"field": source_field, "level": "error", "message": str(exc)})
    raw_status = _text(raw.get("Status"))
    if raw_status is None:
        status = SIOStatus.unassigned
        messages.append(
            {
                "field": "Status",
                "level": "warning",
                "code": "blank_historical_status",
                "message": (
                    "Blank historical status mapped to unassigned; the blank source value is preserved"
                ),
            }
        )
    else:
        try:
            status = _enum_value(raw_status, STATUS_MAP, "Status")
        except ValueError as exc:
            messages.append({"field": "Status", "level": "error", "message": str(exc)})

    for source_field, mapping, target in (
        ("Nature of Observation", NATURE_MAP, "nature"),
        ("Urgency", URGENCY_MAP, "urgency"),
    ):
        try:
            value = _enum_value(raw.get(source_field), mapping, source_field, optional=source_field == "Urgency")
            if target == "nature":
                nature = value
            else:
                urgency = value
        except ValueError as exc:
            messages.append({"field": source_field, "level": "error", "message": str(exc)})

    officer_name = _preserved_text(raw.get("Responsible H&S Officer"))
    responsible_name = _preserved_text(raw.get("Person Responsible for Corrective Action"))
    department_name = _preserved_text(raw.get("Department"))
    responsible_department_name = _preserved_text(
        raw.get("Department Responsible for Corrective Action")
    )
    department_id, department_match = _exact_department_match(db, _text(department_name))
    responsible_department_id, responsible_department_match = _exact_department_match(
        db, _text(responsible_department_name)
    )
    officer_id, officer_match = _exact_user_match(db, _text(officer_name))
    responsible_person_id, responsible_person_match = _exact_user_match(
        db, _text(responsible_name)
    )
    for mapping_message in (
        _mapping_message("Department", department_name, department_match),
        _mapping_message(
            "Department Responsible for Corrective Action",
            responsible_department_name,
            responsible_department_match,
        ),
        _mapping_message("Responsible H&S Officer", officer_name, officer_match),
        _mapping_message(
            "Person Responsible for Corrective Action",
            responsible_name,
            responsible_person_match,
        ),
    ):
        if mapping_message:
            messages.append(mapping_message)

    unexpected_source_fields = {
        key: _json_value(value)
        for key, value in raw.items()
        if key not in YALELO_SIO_COLUMNS and _text(value) is not None
    }
    legacy_metadata = {
        "MonthY": _json_value(raw.get("MonthY")),
        "Item Type": _json_value(raw.get("Item Type")),
        "source_status": _json_value(raw.get("Status")),
        "source_observation_nature": _json_value(raw.get("Nature of Observation")),
        "source_urgency": _json_value(raw.get("Urgency")),
    }
    if unexpected_source_fields:
        legacy_metadata["unexpected_source_fields"] = unexpected_source_fields
    normalized = {
        "external_reference_id": _text(raw.get("ID")),
        "source_system": YALELO_SOURCE_SYSTEM,
        "observation_date": observation_date.isoformat() if observation_date else None,
        "department": department_name,
        "department_id": department_id,
        "source_type": _preserved_text(raw.get("Source of Observation")),
        "description": _preserved_text(raw.get("Description of SIO")),
        "incident_classification": _preserved_text(raw.get("Incident Classification")),
        "status": status.value if status else None,
        "observation_nature": nature.value if nature else None,
        "responsible_department": responsible_department_name,
        "responsible_department_id": responsible_department_id,
        "responsible_hs_officer_user_id": officer_id,
        "responsible_hs_officer_name": officer_name,
        "urgency": urgency.value if urgency else None,
        "category": _preserved_text(raw.get("SIO Category")),
        "responsible_person_user_id": responsible_person_id,
        "responsible_user_id": responsible_person_id,
        "responsible_person_name": responsible_name,
        "property_damage": _preserved_text(raw.get("Property Damage")),
        "source_created_at": source_created_at.isoformat() if source_created_at else None,
        "source_created_by": _preserved_text(raw.get("Created By")),
        "source_modified_by": _preserved_text(raw.get("Modified By")),
        "source_path": _preserved_text(raw.get("Path")),
        "legacy_metadata": legacy_metadata,
    }
    return normalized, messages


def preview_yalelo_sio_import(
    db: Session,
    *,
    content: bytes,
    filename: str,
    actor_id: Optional[int],
) -> DataImportJob:
    if not filename.lower().endswith(".xlsx"):
        raise ImportValidationError("Only .xlsx workbooks are supported")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError("The uploaded file is not a readable .xlsx workbook") from exc
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration as exc:
            raise ImportValidationError("The workbook is empty") from exc
        headers = [_text(value) for value in header_values]
        duplicate_columns = sorted(
            column for column, count in Counter(headers).items() if column is not None and count > 1
        )
        if duplicate_columns:
            raise ImportValidationError(
                "Workbook contains duplicate columns: " + ", ".join(duplicate_columns)
            )
        if any(header is None for header in headers):
            raise ImportValidationError("Workbook contains one or more unnamed columns")
        missing_columns = [column for column in YALELO_SIO_COLUMNS if column not in headers]
        if missing_columns:
            raise ImportValidationError(
                "Workbook does not match the Yalelo SIO format; missing columns: "
                + ", ".join(missing_columns)
            )
        additional_columns = [column for column in headers if column not in YALELO_SIO_COLUMNS]

        job = DataImportJob(
            importer_type=YALELO_SIO_IMPORTER,
            source_system=YALELO_SOURCE_SYSTEM,
            original_filename=filename,
            status=ImportJobStatus.previewed,
            is_dry_run=True,
            started_at=_now(),
            created_by_user_id=actor_id,
        )
        db.add(job)
        db.flush()
        site_by_name, ambiguous_site_names = _site_lookup(db)
        seen_external_ids: set[str] = set()
        unresolved_sites: set[str] = set()
        site_mappings: dict[str, int] = {}
        validation_messages: list[dict] = []
        if additional_columns:
            validation_messages.append(
                {
                    "level": "warning",
                    "field": "Workbook",
                    "code": "additional_columns",
                    "message": "Additional columns are preserved in legacy metadata: "
                    + ", ".join(additional_columns),
                }
            )
        detected = valid = duplicates = failed = 0
        populated_columns: set[str] = set()
        unresolved_departments: set[str] = set()
        unresolved_responsible_departments: set[str] = set()
        resolved_users: set[str] = set()
        unresolved_users: set[str] = set()
        ambiguous_users: set[str] = set()
        unexpected_statuses: set[str] = set()
        unexpected_urgencies: set[str] = set()
        unexpected_classifications: set[str] = set()
        malformed_dates: list[dict] = []

        for row_number, values in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            detected += 1
            raw = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header is not None
            }
            populated_columns.update(
                key for key, value in raw.items() if _text(value) is not None
            )
            normalized, messages = _normalize_yalelo_row(db, raw, epoch=workbook.epoch)
            if normalized.get("department") and normalized.get("department_id") is None:
                unresolved_departments.add(normalized["department"])
            if (
                normalized.get("responsible_department")
                and normalized.get("responsible_department_id") is None
            ):
                unresolved_responsible_departments.add(normalized["responsible_department"])
            for name_key, id_key in (
                ("responsible_hs_officer_name", "responsible_hs_officer_user_id"),
                ("responsible_person_name", "responsible_person_user_id"),
            ):
                if normalized.get(name_key) and normalized.get(id_key) is not None:
                    resolved_users.add(normalized[name_key])
            for message in messages:
                field = message.get("field")
                code = message.get("code")
                if field in {"Responsible H&S Officer", "Person Responsible for Corrective Action"}:
                    source_value = message.get("source_value")
                    if source_value and code == "ambiguous_mapping":
                        ambiguous_users.add(source_value)
                    elif source_value and code == "unresolved_mapping":
                        unresolved_users.add(source_value)
                if field == "Status" and message.get("level") == "error":
                    unexpected_statuses.add(_text(raw.get("Status")) or "<blank>")
                if field == "Urgency" and message.get("level") == "error":
                    unexpected_urgencies.add(_text(raw.get("Urgency")) or "<blank>")
                if field in {"Date", "Created"} and message.get("level") == "error":
                    malformed_dates.append(
                        {
                            "row_number": row_number,
                            "field": field,
                            "value": _json_value(raw.get(field)),
                            "message": message.get("message"),
                        }
                    )
            classification = _text(raw.get("Incident Classification"))
            if classification and _key(classification) not in YALELO_KNOWN_INCIDENT_CLASSIFICATIONS:
                unexpected_classifications.add(classification)
            external_id = normalized.get("external_reference_id")
            site_name = _text(raw.get("Site"))
            resolved_site_id = site_by_name.get(site_name.lower()) if site_name else None
            if resolved_site_id is not None:
                site_mappings[site_name] = resolved_site_id
            elif site_name:
                unresolved_sites.add(site_name)
                if site_name.lower() in ambiguous_site_names:
                    messages.append(
                        {
                            "field": "Site",
                            "level": "warning",
                            "message": "Multiple existing sites have this name; select the intended site",
                        }
                    )

            is_duplicate = False
            if external_id:
                is_duplicate = external_id in seen_external_ids or db.scalar(
                    select(SafetyImprovementObservation.id).where(
                        SafetyImprovementObservation.source_system == YALELO_SOURCE_SYSTEM,
                        SafetyImprovementObservation.external_reference_id == external_id,
                    )
                ) is not None
                seen_external_ids.add(external_id)

            errors = [message for message in messages if message.get("level") == "error"]
            if is_duplicate:
                row_status = ImportRowStatus.duplicate
                duplicates += 1
                messages.append(
                    {"field": "ID", "level": "info", "message": "Duplicate source ID will be skipped"}
                )
            elif errors:
                row_status = ImportRowStatus.invalid
                failed += 1
            elif resolved_site_id is None:
                row_status = ImportRowStatus.unresolved_site
                valid += 1
            else:
                row_status = ImportRowStatus.valid
                valid += 1

            row_messages = [dict(message, row_number=row_number) for message in messages]
            validation_messages.extend(row_messages)
            db.add(
                DataImportRow(
                    job_id=job.id,
                    row_number=row_number,
                    external_reference_id=external_id,
                    source_site_name=site_name,
                    resolved_site_id=resolved_site_id,
                    status=row_status,
                    raw_data={key: _json_value(value) for key, value in raw.items()},
                    normalized_data=normalized,
                    messages=row_messages,
                )
            )

        job.total_rows = detected
        job.successful_rows = 0
        job.skipped_rows = duplicates
        job.failed_rows = failed
        job.validation_messages = validation_messages
        completely_empty_columns = [
            column for column in YALELO_SIO_COLUMNS if column not in populated_columns
        ]
        warning_summaries: dict[tuple, dict] = {}
        for message in validation_messages:
            if message.get("level") != "warning":
                continue
            key = (
                message.get("field"),
                message.get("code"),
                message.get("source_value"),
                message.get("message"),
            )
            if key not in warning_summaries:
                warning_summaries[key] = {
                    item_key: item_value
                    for item_key, item_value in message.items()
                    if item_key != "row_number"
                }
                warning_summaries[key]["count"] = 0
            warning_summaries[key]["count"] += 1
        job.report = {
            "rows_detected": detected,
            "rows_valid": valid,
            "rows_imported": 0,
            "duplicates_skipped": duplicates,
            "rows_failed": failed,
            "unresolved_sites": sorted(unresolved_sites),
            "site_mappings": site_mappings,
            "column_contract": {
                "expected_columns": list(YALELO_SIO_COLUMNS),
                "detected_columns": headers,
                "missing_columns": missing_columns,
                "additional_columns": additional_columns,
                "duplicate_columns": duplicate_columns,
                "completely_empty_columns": completely_empty_columns,
            },
            "unresolved_departments": sorted(unresolved_departments),
            "unresolved_responsible_departments": sorted(unresolved_responsible_departments),
            "resolved_users": sorted(resolved_users),
            "unresolved_users": sorted(unresolved_users),
            "ambiguous_users": sorted(ambiguous_users),
            "unexpected_statuses": sorted(unexpected_statuses),
            "unexpected_urgency_values": sorted(unexpected_urgencies),
            "unexpected_classifications": sorted(unexpected_classifications),
            "malformed_dates": malformed_dates,
            "other_warnings": list(warning_summaries.values()),
            "failure_reasons": [message for message in validation_messages if message["level"] == "error"],
        }
        db.add(job)
        db.commit()
        db.refresh(job)
        write_audit_log(
            db,
            actor_id=actor_id,
            action="data_import.preview",
            resource_type="data_import_job",
            resource_id=job.id,
            details={"importer_type": YALELO_SIO_IMPORTER, "rows_detected": detected},
        )
        return job
    finally:
        workbook.close()


def list_import_jobs(db: Session, *, skip: int = 0, limit: int = 50) -> dict:
    statement = select(DataImportJob).order_by(DataImportJob.created_at.desc(), DataImportJob.id.desc())
    items, total = paginate(db, statement, skip=skip, limit=limit)
    return {"items": items, "total": total, "skip": skip, "limit": limit}


def get_import_job(db: Session, job_id: int) -> DataImportJob:
    job = db.get(DataImportJob, job_id)
    if job is None:
        raise ImportJobNotFoundError(f"Import job {job_id} was not found")
    return job


def _unique_site_code(db: Session, name: str) -> str:
    base = re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")[:42] or "SITE"
    base = f"YALELO-{base}"[:50]
    candidate = base
    suffix = 2
    while db.scalar(select(Site.id).where(Site.code == candidate)) is not None:
        candidate = f"{base[:46]}-{suffix}"
        suffix += 1
    return candidate


def _resolve_confirmation_sites(
    db: Session,
    request: ImportConfirmRequest,
    *,
    actor_id: Optional[int],
) -> dict[str, int]:
    mappings: dict[str, int] = {}
    for source_name, site_id in request.site_mappings.items():
        site = db.get(Site, site_id)
        if site is None:
            raise ImportValidationError(f"Mapped site {site_id} for {source_name} was not found")
        mappings[source_name.strip().lower()] = site.id
    existing, ambiguous = _site_lookup(db)
    for source_name in request.create_sites:
        normalized_name = source_name.strip()
        if not normalized_name:
            continue
        key = normalized_name.lower()
        if key in mappings:
            continue
        if key in existing:
            mappings[key] = existing[key]
            continue
        if key in ambiguous:
            raise ImportValidationError(
                f"Site {normalized_name} is ambiguous; map it to a specific existing site"
            )
        site = Site(
            name=normalized_name,
            code=_unique_site_code(db, normalized_name),
            created_by_id=actor_id,
        )
        db.add(site)
        db.flush()
        mappings[key] = site.id
    return mappings


def _sio_input_from_row(row: DataImportRow, site_id: int) -> SIOCreate:
    data = dict(row.normalized_data)
    data["site_id"] = site_id
    if data.get("observation_date"):
        data["observation_date"] = date.fromisoformat(data["observation_date"])
    if data.get("source_created_at"):
        data["source_created_at"] = datetime.fromisoformat(data["source_created_at"])
    return SIOCreate.model_validate(data)


def confirm_yalelo_sio_import(
    db: Session,
    job: DataImportJob,
    request: ImportConfirmRequest,
    *,
    actor_id: Optional[int],
) -> DataImportJob:
    if job.importer_type != YALELO_SIO_IMPORTER:
        raise ImportStateError("The selected import job is not a Yalelo SIO import")
    if job.status in {ImportJobStatus.completed, ImportJobStatus.completed_with_errors}:
        return job
    if job.status != ImportJobStatus.previewed:
        raise ImportStateError(f"Import job cannot be confirmed while it is {job.status.value}")

    preview_report = dict(job.report or {})
    confirmation_mappings = _resolve_confirmation_sites(db, request, actor_id=actor_id)
    job.status = ImportJobStatus.processing
    job.is_dry_run = False
    db.add(job)
    db.commit()

    imported = 0
    duplicates = 0
    failed = 0
    failure_reasons: list[dict] = []
    rows_since_commit = 0
    for row in sorted(job.rows, key=lambda item: item.row_number):
        if row.status == ImportRowStatus.invalid:
            failed += 1
            failure_reasons.extend(
                message for message in row.messages if message.get("level") == "error"
            )
            continue
        if row.status == ImportRowStatus.duplicate:
            duplicates += 1
            continue
        if row.status == ImportRowStatus.imported:
            imported += 1
            continue

        site_id = row.resolved_site_id
        if site_id is None and row.source_site_name:
            site_id = confirmation_mappings.get(row.source_site_name.strip().lower())
        if site_id is None:
            row.status = ImportRowStatus.failed
            row.failure_reason = f"Unresolved site: {row.source_site_name or 'blank'}"
            failure = {
                "row_number": row.row_number,
                "field": "Site",
                "level": "error",
                "message": row.failure_reason,
            }
            row.messages = [*row.messages, failure]
            failure_reasons.append(failure)
            failed += 1
            db.add(row)
            continue

        try:
            with db.begin_nested():
                sio = create_sio(
                    db,
                    _sio_input_from_row(row, site_id),
                    actor_id=actor_id,
                    is_import=True,
                    commit=False,
                )
                row.resolved_site_id = site_id
                row.imported_sio_id = sio.id
                row.status = ImportRowStatus.imported
                row.failure_reason = None
                db.add(row)
                db.flush()
            imported += 1
        except SIODuplicateError:
            row.status = ImportRowStatus.duplicate
            row.failure_reason = None
            duplicates += 1
        except Exception as exc:
            row.status = ImportRowStatus.failed
            row.failure_reason = str(exc)
            failure = {
                "row_number": row.row_number,
                "level": "error",
                "message": str(exc),
            }
            row.messages = [*row.messages, failure]
            failure_reasons.append(failure)
            failed += 1
        db.add(row)
        rows_since_commit += 1
        if rows_since_commit >= 100:
            db.commit()
            rows_since_commit = 0

    if rows_since_commit:
        db.commit()

    job.successful_rows = imported
    job.skipped_rows = duplicates
    job.failed_rows = failed
    job.completed_at = _now()
    job.status = ImportJobStatus.completed_with_errors if failed else ImportJobStatus.completed
    job.report = {
        "rows_detected": job.total_rows,
        "rows_valid": max(0, job.total_rows - failed - duplicates),
        "rows_imported": imported,
        "duplicates_skipped": duplicates,
        "rows_failed": failed,
        "unresolved_sites": sorted(
            {
                row.source_site_name
                for row in job.rows
                if row.status == ImportRowStatus.failed and row.source_site_name
            }
        ),
        "site_mappings": {
            row.source_site_name: row.resolved_site_id
            for row in job.rows
            if row.source_site_name and row.resolved_site_id
        },
        "failure_reasons": failure_reasons,
        **{
            key: preview_report[key]
            for key in (
                "column_contract",
                "unresolved_departments",
                "unresolved_responsible_departments",
                "resolved_users",
                "unresolved_users",
                "ambiguous_users",
                "unexpected_statuses",
                "unexpected_urgency_values",
                "unexpected_classifications",
                "malformed_dates",
                "other_warnings",
            )
            if key in preview_report
        },
    }
    job.validation_messages = [*job.validation_messages, *failure_reasons]
    db.add(job)
    db.commit()
    db.refresh(job)
    write_audit_log(
        db,
        actor_id=actor_id,
        action="data_import.confirm",
        resource_type="data_import_job",
        resource_id=job.id,
        details={"rows_imported": imported, "duplicates_skipped": duplicates, "rows_failed": failed},
    )
    return job
